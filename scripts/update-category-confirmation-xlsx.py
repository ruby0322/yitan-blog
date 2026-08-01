#!/usr/bin/env python3
"""Append new posts to 文章分類確認.xlsx for client category review."""

from __future__ import annotations

import json
from copy import copy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
POSTS_PATH = ROOT / "scripts/seed-data/client-posts.json"
DEFAULT_INPUT = Path.home() / "Downloads/文章分類確認.xlsx"
OUTPUT_PATH = ROOT / "materials/文章分類確認.xlsx"

SHEET_NAME = "文章分類確認"
EXISTING_SLUGS = {
    "tigar-o-chronic-pancreatitis-causes",
    "brca2-hereditary-cancer-pancreatic-risk",
    "is-your-pancreas-aging-early",
    "stage0-micro-pancreatic-cancer-under-1cm",
    "pancreas-three-questions",
    "pancreas-head-neck-body-tail-imaging-report-terms",
    "autoimmune-pancreatitis-aip-mimic-cancer",
    "idiopathic-pancreatitis-genetic-truth",
    "pancreas-diet-fat-or-no-fat",
    "pancreatic-cancer-screening-steps",
    "pancreas-anti-aging-rejuvenation-key",
    "daraxonrasib-pancreatic-cancer",
    "pancreatic-cancer-early-symptoms-6-warning-signs",
    "who-should-watch-pancreas-health",
    "osteoporosis-vitamin-d-pancreatic-function",
}

# Slug column may be blank for legacy rows; map title -> slug for backfill.
TITLE_SLUG_BACKFILL = {
    "你的胰臟初老了嗎？": "is-your-pancreas-aging-early",
}


def load_posts() -> list[dict]:
    with POSTS_PATH.open(encoding="utf-8") as handle:
        return json.load(handle)


def cover_filename(post: dict) -> str | None:
    covers = post.get("covers") or []
    return covers[0] if covers else None


def build_row(post: dict, seq: int) -> list:
    cover = cover_filename(post)
    return [
        seq,
        post["title"],
        post.get("slug"),
        post.get("seoTitle"),
        post.get("metaDescription"),
        post["title"],
        ", ".join(post.get("categories") or []),
        None,  # 確認分類 — client fills in
        None,  # 客戶備註
        len(post.get("covers") or []),
        len(post.get("inlines") or []),
        cover,
        post.get("folder"),
        None,  # 系統備註
    ]


def copy_row_style(ws, template_row: int, target_row: int, max_col: int = 14) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(template_row, col)
        dst = ws.cell(target_row, col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def backfill_missing_slug(ws, posts_by_slug: dict[str, dict]) -> None:
    for row in range(2, ws.max_row + 1):
        slug = ws.cell(row, 3).value
        if slug:
            continue
        title = ws.cell(row, 2).value or ws.cell(row, 6).value
        slug = TITLE_SLUG_BACKFILL.get(title)
        if not slug:
            continue
        post = posts_by_slug[slug]
        ws.cell(row, 3).value = post["slug"]
        if not ws.cell(row, 4).value:
            ws.cell(row, 4).value = post.get("seoTitle")
        if not ws.cell(row, 5).value:
            ws.cell(row, 5).value = post.get("metaDescription")


def update_category_dropdown(ws) -> None:
    categories = "基礎知識,胰臟癌,胰臟發炎,胰臟水泡,飲食保健,健檢判讀,胰臟癌篩檢,胰臟功能,胰臟健康"
    ws.data_validations.dataValidation.clear()
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1=f'"{categories}"', allow_blank=True)
    dv.error = "請從下拉選單選擇分類"
    dv.errorTitle = "分類格式"
    dv.prompt = "可選一個或多個分類（多個請用逗號分隔）"
    dv.promptTitle = "確認分類"
    dv.add(f"H2:H{ws.max_row}")
    ws.add_data_validation(dv)


def main(input_path: Path = DEFAULT_INPUT) -> None:
    posts = load_posts()
    posts_by_slug = {post["slug"]: post for post in posts}
    new_posts = [post for post in posts if post["slug"] not in EXISTING_SLUGS]

    wb = openpyxl.load_workbook(input_path)
    ws = wb[SHEET_NAME]

    backfill_missing_slug(ws, posts_by_slug)

    next_row = 2
    while ws.cell(next_row, 1).value is not None:
        next_row += 1

    start_seq = int(ws.cell(next_row - 1, 1).value or 0) + 1

    for index, post in enumerate(new_posts):
        row = next_row + index
        values = build_row(post, start_seq + index)
        for col, value in enumerate(values, start=1):
            ws.cell(row, col).value = value
        copy_row_style(ws, 2, row)

    update_category_dropdown(ws)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    wb.save(input_path)

    print(f"Updated {len(new_posts)} new posts (rows {start_seq}-{start_seq + len(new_posts) - 1})")
    print(f"Saved: {OUTPUT_PATH}")
    print(f"Saved: {input_path}")


if __name__ == "__main__":
    main()
