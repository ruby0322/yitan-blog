#!/usr/bin/env python3
"""Apply confirmed categories from 文章分類確認.xlsx to client-posts.json."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "materials/文章分類確認.xlsx"
DEFAULT_POSTS = ROOT / "scripts/seed-data/client-posts.json"

SHEET_CONFIRM = "文章分類確認"
SHEET_OPTIONS = "分類選項"

COL_SEQ = 1
COL_TITLE = 2
COL_SLUG = 3
COL_DRAFT = 7
COL_CONFIRMED = 8

TITLE_SLUG_FALLBACK = {
    "你的胰臟初老了嗎？": "is-your-pancreas-aging-early",
}


def load_allowed_categories(xlsx_path: Path) -> set[str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_OPTIONS]
    allowed: set[str] = set()
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row, 1).value
        if title and str(title).strip():
            allowed.add(str(title).strip())
    return allowed


def parse_categories(raw: object) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    parts = re.split(r"[,，]", text)
    seen: set[str] = set()
    result: list[str] = []
    for part in parts:
        label = part.strip()
        if label and label not in seen:
            seen.add(label)
            result.append(label)
    return result


def effective_categories(draft: object, confirmed: object) -> list[str]:
    confirmed_cats = parse_categories(confirmed)
    if confirmed_cats:
        return confirmed_cats
    return parse_categories(draft)


def load_xlsx_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_CONFIRM]
    rows: list[dict] = []
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row, COL_SEQ).value
        if seq is None:
            break
        title = ws.cell(row, COL_TITLE).value
        slug = ws.cell(row, COL_SLUG).value
        draft = ws.cell(row, COL_DRAFT).value
        confirmed = ws.cell(row, COL_CONFIRMED).value
        rows.append(
            {
                "seq": int(seq),
                "title": str(title).strip() if title else "",
                "slug": str(slug).strip() if slug else "",
                "categories": effective_categories(draft, confirmed),
            }
        )
    return rows


def build_category_map(rows: list[dict]) -> dict[str, list[str]]:
    by_slug: dict[str, list[str]] = {}
    by_title: dict[str, list[str]] = {}
    for row in rows:
        if row["slug"]:
            by_slug[row["slug"]] = row["categories"]
        if row["title"]:
            by_title[row["title"]] = row["categories"]
        fallback_slug = TITLE_SLUG_FALLBACK.get(row["title"])
        if fallback_slug and row["categories"]:
            by_slug.setdefault(fallback_slug, row["categories"])
    return {**by_title, **by_slug}


def resolve_categories(post: dict, category_map: dict[str, list[str]]) -> list[str] | None:
    slug = post.get("slug", "")
    if slug and slug in category_map:
        return category_map[slug]
    title = post.get("title", "")
    if title and title in category_map:
        return category_map[title]
    return None


def apply_categories(
    posts_path: Path,
    xlsx_path: Path,
    *,
    dry_run: bool = False,
) -> int:
    allowed = load_allowed_categories(xlsx_path)
    rows = load_xlsx_rows(xlsx_path)
    category_map = build_category_map(rows)

    posts = json.loads(posts_path.read_text(encoding="utf-8"))
    errors: list[str] = []
    changes: list[tuple[str, object, list[str]]] = []

    xlsx_slugs = {row["slug"] for row in rows if row["slug"]}
    xlsx_slugs.update(TITLE_SLUG_FALLBACK.values())
    json_slugs = {post["slug"] for post in posts}

    for slug in sorted(xlsx_slugs - json_slugs):
        errors.append(f"xlsx row slug not found in JSON: {slug}")
    for slug in sorted(json_slugs - xlsx_slugs):
        errors.append(f"JSON slug not found in xlsx: {slug}")

    for post in posts:
        slug = post["slug"]
        new_categories = resolve_categories(post, category_map)
        if new_categories is None:
            errors.append(f"no category mapping for post: {slug}")
            continue
        if not new_categories:
            errors.append(f"empty categories for post: {slug}")
            continue
        for label in new_categories:
            if label not in allowed:
                errors.append(f"unknown category {label!r} for post {slug}")
        old = post.get("categories", post.get("category"))
        if old != new_categories:
            changes.append((slug, old, new_categories))
        post.pop("category", None)
        post["categories"] = new_categories

    if errors:
        for err in errors:
            print(f"ERROR: {err}", file=sys.stderr)
        return 1

    for slug, old, new in changes:
        print(f"{slug}: {old} -> {new}")

    print(f"\n{len(changes)} post(s) changed, {len(posts)} total")

    if not dry_run:
        posts_path.write_text(
            json.dumps(posts, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"Saved: {posts_path}")
    else:
        print("Dry run — no files written")

    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--posts", type=Path, default=DEFAULT_POSTS)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    code = apply_categories(args.posts, args.xlsx, dry_run=args.dry_run)
    raise SystemExit(code)


if __name__ == "__main__":
    main()
