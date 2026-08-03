#!/usr/bin/env python3
"""Generate delivery checklist xlsx for 胰探究竟 project handover."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "docs/delivery-checklist.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="6F8D7A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

DELIVERABLES = [
    # (cat_id, 大類, 項目, 說明, 狀態, 備註, 截圖)
    # A. 品牌官網
    ("A", "品牌官網", "首頁 Hero 主視覺區", "訪客一進網站看到的標題、副標與兩個行動按鈕（閱讀最新文章、認識章醫師）", "已完成", "", "delivery-screenshots/01-home-hero-desktop.png"),
    ("A", "品牌官網", "首頁品牌引言區", "展示「健康，來自理解，而非恐懼」等品牌引言與說明文字", "已完成", "", "delivery-screenshots/01-home-hero-desktop.png"),
    ("A", "品牌官網", "首頁四大特色區", "四格卡片介紹近 30 年臨床經驗、完整疾病光譜、早期發現、理解而非恐懼", "已完成", "", "delivery-screenshots/02-home-features-desktop.png"),
    ("A", "品牌官網", "首頁精選文章區", "可手動挑選 1–3 篇重點文章，引導讀者深入閱讀", "已完成", "", "delivery-screenshots/07-post-detail-desktop.png"),
    ("A", "品牌官網", "首頁依主題閱讀區", "五個主題分類卡片（胰臟水泡、發炎、癌、篩檢、健康），可點進各主題", "已完成", "", "delivery-screenshots/03-home-categories-desktop.png"),
    ("A", "品牌官網", "首頁關於章醫師區", "簡介章醫師學經歷與專業背景，連結至關於頁", "已完成", "", "delivery-screenshots/04-home-book-desktop.png"),
    ("A", "品牌官網", "首頁書籍推廣區", "《攔截胰臟癌》書籍介紹與博客來購買連結", "已完成", "", "delivery-screenshots/04-home-book-desktop.png"),
    ("A", "品牌官網", "關於頁", "完整介紹章醫師理念、專業經歷、網站成立目的與醫療資訊聲明", "已完成", "", "delivery-screenshots/05-about-desktop.png"),
    ("A", "品牌官網", "網站導覽列", "頂部選單連結，可在後台調整項目與連結", "已完成", "", "delivery-screenshots/01-home-hero-desktop.png"),
    ("A", "品牌官網", "網站頁尾", "底部連結群組與深綠色品牌頁尾", "已完成", "", "delivery-screenshots/04-home-book-desktop.png"),
    ("A", "品牌官網", "手機版版面", "手機、平板、電腦螢幕皆可正常瀏覽，選單改為漢堡按鈕", "已完成", "", "delivery-screenshots/09-home-mobile.png"),
    ("A", "品牌官網", "網站 Logo 與品牌圖形", "Logo 與胰臟品牌圖形由設計固定，無法在後台自行更換", "已完成", "需工程師修改", "delivery-screenshots/01-home-hero-desktop.png"),
    # B. 文章閱讀
    ("B", "文章閱讀", "文章列表頁", "所有已發布文章以卡片形式列出，顯示日期、分類、摘要", "已完成", "", "delivery-screenshots/06-posts-list-desktop.png"),
    ("B", "文章閱讀", "分類篩選", "可依五個主題分類篩選文章（網址帶 ?category= 參數）", "已完成", "", "delivery-screenshots/06-posts-list-desktop.png"),
    ("B", "文章閱讀", "分頁瀏覽", "文章超過 12 篇時自動分頁", "已完成", "", "delivery-screenshots/06-posts-list-desktop.png"),
    ("B", "文章閱讀", "單篇文章頁面", "含封面圖、標題、摘要、正文、發布日期、分類標籤", "已完成", "", "delivery-screenshots/07-post-detail-desktop.png"),
    ("B", "文章閱讀", "文章 FAQ 區塊", "每篇文章可設定常見問答，以摺疊方式呈現", "已完成", "", "delivery-screenshots/07-post-detail-desktop.png"),
    ("B", "文章閱讀", "相關文章推薦", "文章底部可手動指定相關文章連結", "已完成", "", "delivery-screenshots/07-post-detail-desktop.png"),
    ("B", "文章閱讀", "文章法遵聲明", "每篇文章底部顯示醫療資訊免責聲明", "已完成", "文案需工程師修改", "delivery-screenshots/11-terms-desktop.png"),
    # C. 搜尋
    ("C", "搜尋功能", "全站搜尋按鈕", "每頁頂部有搜尋圖示，點擊開啟搜尋視窗", "已完成", "", "delivery-screenshots/08-search-overlay-desktop.png"),
    ("C", "搜尋功能", "即時搜尋結果", "輸入關鍵字後即時顯示符合的文章標題與摘要", "已完成", "", "delivery-screenshots/08-search-overlay-desktop.png"),
    ("C", "搜尋功能", "搜尋結果跳轉", "點選結果直接進入文章；也可查看全部結果", "已完成", "", "delivery-screenshots/08-search-overlay-desktop.png"),
    # D. 內容管理後台
    ("D", "內容管理後台", "後台登入", "繁體中文管理介面，網址 /admin", "已完成", "", "delivery-screenshots/12-admin-login.png"),
    ("D", "內容管理後台", "新增與編輯文章", "可撰寫標題、摘要、正文、上傳封面圖", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("D", "內容管理後台", "文章內插入圖片", "正文編輯器可插入圖片，或從媒體庫選取", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("D", "內容管理後台", "文章內插入表格", "正文編輯器支援表格（實驗性功能）", "已完成", "若異常請改以條列代替", "delivery-screenshots/14-admin-post-edit.png"),
    ("D", "內容管理後台", "草稿儲存", "撰寫中可先存草稿，訪客看不到", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("D", "內容管理後台", "預覽功能", "發布前可預覽文章在網站上的實際效果", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("D", "內容管理後台", "排程發布（後台 UI）", "可在發布按鈕選單設定未來發布/下架日期與時間", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("D", "內容管理後台", "排程自動執行（正式環境）", "到點自動發布需 Vercel Cron 定期呼叫 /api/payload-jobs/run；目前 vercel.json 尚未設定 crons", "部分完成", "需工程師在 vercel.json 加 cron + 確認 CRON_SECRET", ""),
    ("D", "內容管理後台", "媒體庫管理", "集中管理上傳的圖片，可填替代文字與說明", "已完成", "", "delivery-screenshots/16-admin-media.png"),
    # E. 網站設定（後台可改）
    ("E", "網站設定", "編輯首頁各區文案", "後台「頁面 → 首頁」可修改 Hero、引言、特色、精選文章等區塊", "已完成", "", "delivery-screenshots/15-admin-page-home.png"),
    ("E", "網站設定", "編輯關於頁", "後台「頁面 → 關於」可修改各段落內容", "已完成", "", "delivery-screenshots/15-admin-page-home.png"),
    ("E", "網站設定", "調整導覽列連結", "後台 Globals → Header 可增刪改最多 6 個導覽項目", "已完成", "", "delivery-screenshots/13-admin-dashboard.png"),
    ("E", "網站設定", "調整頁尾連結", "後台 Globals → Footer 可編輯最多 4 組連結群組", "已完成", "", "delivery-screenshots/13-admin-dashboard.png"),
    ("E", "網站設定", "文章分類管理", "後台「分類」可編輯五個主題名稱、說明與排序", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("E", "網站設定", "每篇文章 SEO 設定", "可自訂搜尋引擎標題、描述、分享預覽圖", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("E", "網站設定", "行銷素材欄位", "後台可記錄 YouTube、社群貼文、電子報摘要等文案（僅後台可見）", "已完成", "", "delivery-screenshots/14-admin-post-edit.png"),
    # F. 內容資料
    ("F", "內容資料", "五大主題分類", "胰臟水泡、胰臟發炎、胰臟癌、胰臟癌篩檢、胰臟健康", "已完成", "", "delivery-screenshots/03-home-categories-desktop.png"),
    ("F", "內容資料", "29 篇客戶文章匯入", "依提供之 Word 與圖片已匯入後台（含封面與內文配圖）", "部分完成", "多數為草稿，待審核後發布", "delivery-screenshots/14-admin-post-edit.png"),
    ("F", "內容資料", "1 篇示範文章已公開", "「章醫師走過的胰臟癌治療三十年」已上線供展示", "已完成", "", "delivery-screenshots/07-post-detail-desktop.png"),
    # G. 上線與維運
    ("G", "上線與維運", "正式網域上線", "https://pancreasblog.com 已部署上線", "已完成", "", "delivery-screenshots/01-home-hero-desktop.png"),
    ("G", "上線與維運", "排程發布 Cron 觸發", "Vercel Cron 每 5 分鐘呼叫 /api/payload-jobs/run，執行已到期的排程 job", "部分完成", "payload.config 已設 CRON_SECRET 驗證；vercel.json 待加 crons", ""),
    ("G", "上線與維運", "雲端資料庫", "文章與設定儲存於 Vercel Postgres", "已完成", "", ""),
    ("G", "上線與維運", "雲端圖片儲存", "上傳圖片儲存於 Vercel Blob", "已完成", "", "delivery-screenshots/16-admin-media.png"),
    ("G", "上線與維運", "網站地圖與搜尋引擎", "自動產生 sitemap.xml 與 robots.txt", "已完成", "", ""),
    ("G", "上線與維運", "操作手冊", "後台操作與交接文件", "已完成", "", "delivery-screenshots/13-admin-dashboard.png"),
    ("G", "上線與維運", "電子報訂閱功能", "首頁有訂閱表單外觀，但尚未串接寄信服務", "部分完成", "需工程師後續建置", "delivery-screenshots/04-home-book-desktop.png"),
    ("G", "上線與維運", "法遵與著作權文案", "著作權、商標、醫療聲明等固定文案", "已完成", "需工程師修改，非後台按鈕", "delivery-screenshots/11-terms-desktop.png"),
]

DEMO_AGENDA = [
    (1, "開場", "—", "說明今日走查流程與網站定位", "介紹「胰探究竟」是章醫師的胰臟健康衛教部落格", 3, "", "可先展示首頁截圖預覽全貌"),
    (2, "首頁 Hero + 品牌引言", "https://pancreasblog.com/", "開啟首頁，展示主標「看懂胰臟，從理解開始」與兩個 CTA 按鈕", "強調品牌定位：把艱深醫學說成聽得懂的話", 5, "delivery-screenshots/01-home-hero-desktop.png", "向下捲動可看到品牌引言 Quote 區"),
    (3, "四大特色", "https://pancreasblog.com/", "捲動至「四大特色」區（編號 01）", "近 30 年臨床、完整疾病光譜、早期發現、理解而非恐懼", 3, "delivery-screenshots/02-home-features-desktop.png", ""),
    (4, "精選文章", "https://pancreasblog.com/", "展示「精選文章」區（編號 02），點擊示範文章", "可說明此區可在後台手動挑選重點文章", 3, "delivery-screenshots/07-post-detail-desktop.png", ""),
    (5, "依主題閱讀", "https://pancreasblog.com/", "展示五個主題分類卡片（編號 03）", "五個主題：水泡、發炎、癌、篩檢、健康", 5, "delivery-screenshots/03-home-categories-desktop.png", "部分主題尚無已發布文章"),
    (6, "關於章醫師 + 書籍", "https://pancreasblog.com/about", "首頁 About 區 + 書籍區，再進入關於頁", "章醫師學經歷、《攔截胰臟癌》書籍推廣", 7, "delivery-screenshots/04-home-book-desktop.png\ndelivery-screenshots/05-about-desktop.png", ""),
    (7, "文章列表", "https://pancreasblog.com/posts", "展示文章列表與分類篩選", "目前 1 篇已發布，其餘 29 篇在後台草稿待審", 5, "delivery-screenshots/06-posts-list-desktop.png", "可點分類標籤示範篩選"),
    (8, "單篇文章", "https://pancreasblog.com/posts/pancreatic-cancer-treatment-30-years", "展示完整文章：封面、內文、FAQ、法遵聲明", "示範文章含圖文、表格、FAQ 摺疊區", 8, "delivery-screenshots/07-post-detail-desktop.png\ndelivery-screenshots/10-post-mobile.png", ""),
    (9, "搜尋", "https://pancreasblog.com/", "點頂部搜尋按鈕 → 輸入「胰臟」→ 點結果", "全站即時搜尋，方便讀者找文章", 5, "delivery-screenshots/08-search-overlay-desktop.png", "建議搜尋「胰臟癌」或「三十年」"),
    (10, "手機版", "https://pancreasblog.com/", "縮小瀏覽器視窗或使用手機展示", "確認手機版可讀、選單可用", 3, "delivery-screenshots/09-home-mobile.png\ndelivery-screenshots/10-post-mobile.png", ""),
    (11, "法遵頁面", "https://pancreasblog.com/terms", "展示服務條款與醫療資訊聲明", "著作權、商標、免責聲明", 2, "delivery-screenshots/11-terms-desktop.png", ""),
    (12, "後台登入", "https://pancreasblog.com/admin", "登入繁中後台，展示儀表板", "說明這是章醫師自行發文的管理後台", 3, "delivery-screenshots/12-admin-login.png\ndelivery-screenshots/13-admin-dashboard.png", ""),
    (13, "新增/編輯文章", "https://pancreasblog.com/admin/collections/posts", "開啟一篇文章或建立新項目，展示各欄位", "標題、摘要、內文、封面、分類、FAQ、行銷素材", 10, "delivery-screenshots/14-admin-post-edit.png", "勿在正式環境點「建立示範內容」"),
    (14, "預覽與發布", "https://pancreasblog.com/admin", "示範草稿 → 預覽 → 發布流程", "建議流程：撰寫 → 存草稿 → 預覽 → 發布", 5, "delivery-screenshots/14-admin-post-edit.png", ""),
    (15, "編輯首頁區塊", "https://pancreasblog.com/admin/collections/pages", "開啟「首頁」，示範修改一段文案", "首頁各區塊皆可在後台編輯", 5, "delivery-screenshots/15-admin-page-home.png", ""),
    (16, "媒體庫", "https://pancreasblog.com/admin/collections/media", "展示已上傳圖片與管理介面", "上傳、替代文字、說明", 3, "delivery-screenshots/16-admin-media.png", ""),
    (17, "Q&A 收尾", "—", "說明後續發文節奏與需工程師協助的項目", "29 篇草稿待審、排程發布待加 Vercel Cron、電子報未串接、法遵文案需工程師改", 5, "", ""),
]

SCREENSHOTS = [
    ("01-home-hero-desktop.png", "首頁 Hero 與品牌引言", "前台", "https://pancreasblog.com/"),
    ("02-home-features-desktop.png", "首頁四大特色", "前台", "https://pancreasblog.com/"),
    ("03-home-categories-desktop.png", "首頁依主題閱讀", "前台", "https://pancreasblog.com/"),
    ("04-home-book-desktop.png", "首頁書籍推廣區", "前台", "https://pancreasblog.com/"),
    ("05-about-desktop.png", "關於頁", "前台", "https://pancreasblog.com/about"),
    ("06-posts-list-desktop.png", "文章列表", "前台", "https://pancreasblog.com/posts"),
    ("07-post-detail-desktop.png", "示範文章內頁", "前台", "https://pancreasblog.com/posts/pancreatic-cancer-treatment-30-years"),
    ("08-search-overlay-desktop.png", "搜尋彈窗", "前台", "https://pancreasblog.com/"),
    ("09-home-mobile.png", "首頁手機版", "前台", "https://pancreasblog.com/"),
    ("10-post-mobile.png", "文章手機版", "前台", "https://pancreasblog.com/posts/pancreatic-cancer-treatment-30-years"),
    ("11-terms-desktop.png", "法遵頁", "前台", "https://pancreasblog.com/terms"),
    ("12-admin-login.png", "後台登入", "後台", "https://pancreasblog.com/admin/login"),
    ("13-admin-dashboard.png", "後台儀表板", "後台", "https://pancreasblog.com/admin"),
    ("14-admin-post-edit.png", "文章編輯", "後台", "https://pancreasblog.com/admin/collections/posts"),
    ("15-admin-page-home.png", "首頁區塊編輯", "後台", "https://pancreasblog.com/admin/collections/pages"),
    ("16-admin-media.png", "媒體庫", "後台", "https://pancreasblog.com/admin/collections/media"),
]

ADMIN_EDITABLE = [
    ("文章標題、摘要、正文", "是", "文章 → 建立/編輯", "否", "日常發文主要操作", "delivery-screenshots/14-admin-post-edit.png"),
    ("文章封面圖", "是", "文章 → 內容分頁 → 封面圖", "否", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("文章分類", "是", "文章 → 其他分頁 → 分類", "否", "可選多個主題", "delivery-screenshots/14-admin-post-edit.png"),
    ("文章 FAQ", "是", "文章 → 其他分頁 → FAQ", "否", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("相關文章", "是", "文章 → 其他分頁 → 相關文章", "否", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("行銷素材（YouTube/社群文案）", "是", "文章 → 其他分頁 → 行銷素材", "否", "僅後台可見，不顯示於前台", "delivery-screenshots/14-admin-post-edit.png"),
    ("文章 SEO（標題/描述/分享圖）", "是", "文章 → SEO 分頁", "否", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("草稿 / 預覽", "是", "文章編輯頁右上角狀態", "否", "", "delivery-screenshots/14-admin-post-edit.png"),
    ("排程發布（設定時間）", "是", "發布按鈕 → 排程發布", "否", "後台可排程，job 寫入 payload-jobs", "delivery-screenshots/14-admin-post-edit.png"),
    ("排程到點自動執行", "否", "—", "是", "需 vercel.json crons 呼叫 /api/payload-jobs/run", ""),
    ("首頁 Hero 標題與按鈕", "是", "頁面 → 首頁 → Hero 區塊", "否", "", "delivery-screenshots/15-admin-page-home.png"),
    ("首頁各內容區塊（引言、特色、精選等）", "是", "頁面 → 首頁 → 內容分頁", "否", "共 6 個可編輯區塊", "delivery-screenshots/15-admin-page-home.png"),
    ("關於頁內容", "是", "頁面 → 關於", "否", "", "delivery-screenshots/05-about-desktop.png"),
    ("導覽列連結", "是", "Globals → Header", "否", "最多 6 項", "delivery-screenshots/13-admin-dashboard.png"),
    ("頁尾連結群組", "是", "Globals → Footer", "否", "最多 4 組", "delivery-screenshots/13-admin-dashboard.png"),
    ("文章分類名稱與排序", "是", "分類", "否", "sortOrder 數字越小越前面", "delivery-screenshots/14-admin-post-edit.png"),
    ("媒體庫圖片", "是", "媒體庫", "否", "含替代文字與說明", "delivery-screenshots/16-admin-media.png"),
    ("後台使用者帳號", "是", "Users", "否", "目前所有 admin 權限相同", "delivery-screenshots/13-admin-dashboard.png"),
    ("網站 Logo", "否", "—", "是", "src/components/Logo/", "delivery-screenshots/01-home-hero-desktop.png"),
    ("品牌胰臟圖形（Hero 右側）", "否", "—", "是", "public/elegant-pancreas.PNG", "delivery-screenshots/01-home-hero-desktop.png"),
    ("網站名稱、標語、關鍵字", "否", "—", "是", "src/constants/site.ts", ""),
    ("著作權 / 商標 / 醫療聲明文案", "否", "—", "是", "src/content/legal.ts", "delivery-screenshots/11-terms-desktop.png"),
    ("服務條款頁內容", "否", "—", "是", "src/app/(frontend)/terms/", "delivery-screenshots/11-terms-desktop.png"),
    ("電子報寄信功能", "否", "—", "是", "前台有表單，尚未串接", "delivery-screenshots/04-home-book-desktop.png"),
    ("網址重新導向規則", "否", "—", "是", "後台隱藏，需工程師操作", ""),
    ("整體配色與字型", "否", "—", "是", "主題 CSS，非後台設定", "delivery-screenshots/01-home-hero-desktop.png"),
]


def style_header_row(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def auto_width(ws, min_width: int = 12, max_width: int = 48) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def write_deliverables(ws) -> None:
    headers = ["編號", "大類", "交付項目", "白話說明（給非技術人員）", "狀態", "備註", "參考截圖"]
    style_header_row(ws, headers)
    for row_idx, (_cat_id, cat, item, desc, status, note, screenshot) in enumerate(DELIVERABLES, start=2):
        ws.append([row_idx - 1, cat, item, desc, status, note, screenshot])
    auto_width(ws)


def write_demo_agenda(ws) -> None:
    headers = [
        "順序",
        "環節",
        "網址",
        "操作步驟",
        "展示重點（給非技術聽眾）",
        "預估分鐘",
        "截圖檔名",
        "講者備註",
    ]
    style_header_row(ws, headers)
    for row in DEMO_AGENDA:
        ws.append(list(row))
    total_mins = sum(r[5] for r in DEMO_AGENDA)
    ws.append(["", "合計", "", "", "", total_mins, "", "約 45–60 分鐘含 Q&A"])
    auto_width(ws)


def write_admin_editable(ws) -> None:
    headers = ["區域", "後台能改？", "後台路徑", "需工程師？", "備註", "參考截圖"]
    style_header_row(ws, headers)
    for row in ADMIN_EDITABLE:
        ws.append(list(row))
    auto_width(ws)


def write_screenshots(ws) -> None:
    headers = ["檔名", "說明", "類型", "對應網址", "路徑"]
    style_header_row(ws, headers)
    for filename, desc, kind, url in SCREENSHOTS:
        ws.append([filename, desc, kind, url, f"delivery-screenshots/{filename}"])
    auto_width(ws)


def main() -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "交付項目"
    write_deliverables(ws1)

    ws2 = wb.create_sheet("Demo走查Agenda")
    write_demo_agenda(ws2)

    ws3 = wb.create_sheet("後台可編輯對照")
    write_admin_editable(ws3)

    ws4 = wb.create_sheet("Screenshots")
    write_screenshots(ws4)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
